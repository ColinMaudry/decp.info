import io

import openpyxl
import polars as pl

from src.utils.table import write_styled_excel


def test_write_styled_excel_header_bold_and_color():
    df = pl.DataFrame({"objet": ["marché A"], "montant": [1000.0]})
    buf = io.BytesIO()
    write_styled_excel(df, buf)
    buf.seek(0)
    wb = openpyxl.load_workbook(buf)
    ws = wb.active
    cell = ws.cell(row=1, column=1)
    assert cell.font.bold is True
    assert cell.fill.fgColor.rgb == "FFB33821"
    assert cell.font.color.rgb == "FFFFFFFF"


def test_write_styled_excel_data_cell_text_wrap():
    df = pl.DataFrame({"objet": ["marché A"]})
    buf = io.BytesIO()
    write_styled_excel(df, buf)
    buf.seek(0)
    wb = openpyxl.load_workbook(buf)
    ws = wb.active
    cell = ws.cell(row=2, column=1)
    assert cell.alignment.wrap_text is True


def test_write_styled_excel_known_column_wider_than_minimum():
    # "objet" a une largeur explicite (350px) > minimum (132px)
    # "autre" n'a pas de largeur explicite → reçoit le minimum
    df = pl.DataFrame({"autre": ["x"], "objet": ["y"]})
    buf = io.BytesIO()
    write_styled_excel(df, buf)
    buf.seek(0)
    wb = openpyxl.load_workbook(buf)
    ws = wb.active
    width_autre = ws.column_dimensions["A"].width  # colonne 1 = "autre"
    width_objet = ws.column_dimensions["B"].width  # colonne 2 = "objet"
    assert width_autre > 0
    assert width_objet > width_autre


def test_write_styled_excel_custom_worksheet_name():
    df = pl.DataFrame({"col": ["val"]})
    buf = io.BytesIO()
    write_styled_excel(df, buf, worksheet="2025")
    buf.seek(0)
    wb = openpyxl.load_workbook(buf)
    assert "2025" in wb.sheetnames
