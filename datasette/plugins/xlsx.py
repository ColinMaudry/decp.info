from datasette import hookimpl
from datasette.utils.asgi import Response
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from tempfile import NamedTemporaryFile


def render_spreadsheet(rows):
    wb = Workbook(write_only=True)
    ws = wb.create_sheet()
    ws = wb.active
    ws.title = "decp"

    columnSpecs = {
        'acheteur.nom': {
            'width': 7,
            'wrapText': True
        },
        'objet': {
            'width': 7,
            'wrapText': True
        },
        'titulaire.denominationSociale': {
            'width': 7,
            'wrapText': True
        },
    }

    columns = rows[0].keys()
    if columns[0] == "rowid":
        columns = columns[1:]

    letters = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'

    headers = []
    index = 0
    for col in columns :
        c = WriteOnlyCell(ws, col)
        c.fill = PatternFill("solid", fgColor="DDEFFF")
        headers.append(c)
        if col in columnSpecs:
            ws.column_dimensions[letters[index]].width = columnSpecs[col]['width'] * 5
        else:
            ws.column_dimensions[letters[index]].bestFit = True
        index = index + 1
    ws.append(headers)

    for row in rows:
        wsRow = []
        for col in columns:
            c = WriteOnlyCell(ws, row[col])
            if col in columnSpecs :
                c.alignment = Alignment(wrapText = columnSpecs[col]['wrapText'])
            wsRow.append(c)
        ws.append(wsRow)

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        return Response(
            tmp.read(),
            headers={
                'Content-Disposition': 'attachment; filename=decp.xlsx',
                'Content-type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            }
        )


@hookimpl
def register_output_renderer():
    return {"extension": "xlsx",
    "render": render_spreadsheet,
    "can_render": lambda: False}
